#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FastMoss选品数据解析脚本
读取Excel文件并输出为文本格式
"""

import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("正在安装openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

# 读取Excel文件
excel_file = r"C:\Users\MECHREVO\Downloads\商品新品榜_20260220_210922.xlsx"
wb = load_workbook(excel_file)

print("=" * 80)
print("FastMoss 商品新品榜数据解析")
print("日期：2026-02-20")
print("=" * 80)

# 获取第一个工作表
ws = wb.active
print(f"\n工作表名称: {ws.title}")
print(f"行数: {ws.max_row}")
print(f"列数: {ws.max_column}")
print(f"\n" + "=" * 80 + "\n")

# 读取数据
data = []
header = None

for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row_num == 1:
        header = row  # 第一行是表头
        print("表头：")
        print(" | ".join([str(cell) if cell else "空" for cell in header]))
        print(" | ".join(["-" * 20 for _ in range(20)]))
    else:
        data.append(row)
        # 打印前20行数据（不包括表头）
        if row_num <= 25:
            print(f"行{row_num}: | ".join([str(cell) if cell else "空" for cell in row[:15]]))

# 输出完整数据（所有商品）
print("\n" + "=" * 80)
print("完整数据（所有商品）：")
print("=" * 80)

for i, row in enumerate(data, 1):
    print(f"\n第{i}个商品：")
    try:
        print(f"  排名：{row[0]}")
        print(f"  商品名称：{row[1]}")
        print(f"  价格：{row[2]}")
        print(f"  销量：{row[3]}")
        print(f"  销量环比：{row[4]}")
        print(f"  好评率：{row[5]}")
        print(f"  类目：{row[6]}")
        if len(row) > 7:
            print(f"  SKUs：{row[7]}")
            if len(row) > 8:
                print(f"  其他：{row[8]}")
    except Exception as e:
        print(f"  解析错误：{e}")

print("\n" + "=" * 80)
print("解析完成！")
print("=" * 80)
